# NCRE Excel 操作题私有判分脚本（经典题库第4套：人口普查数据整合，command_metrics 协议）
# 用 openpyxl 检查 Excel.xlsx：两表导入与表格样式、千分位、合并排序、增长数/比重变化、
# 统计指标与降序榜（透视降级扁平汇总）。产物缺失时全 0。
import contextlib
import json
import sys

with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

metrics = {f"e{i}": 0.0 for i in range(1, 9)}
evidence = {}

# 31地区 (2000年人口, 2010年人口)
CENSUS = {
    "北京": (13819000, 19612368), "天津": (10008000, 12938224), "河北": (67440000, 71854202),
    "山西": (32972000, 35712111), "内蒙古": (23760000, 24706321), "辽宁": (42380000, 43746323),
    "吉林": (27280000, 27462297), "黑龙江": (36890000, 38312224), "上海": (16738000, 23019148),
    "江苏": (74383000, 78659903), "浙江": (46769800, 54426891), "安徽": (59860000, 59500510),
    "福建": (34710000, 36894216), "江西": (41400000, 44567475), "山东": (90790000, 95793065),
    "河南": (92560000, 94023567), "湖北": (60278000, 57237740), "湖南": (64400000, 65683722),
    "广东": (86420000, 104303132), "广西": (44890000, 46026629), "海南": (7870000, 8671518),
    "重庆": (30900000, 28846170), "四川": (83290000, 80418200), "贵州": (35250000, 34746468),
    "云南": (42880000, 45966239), "西藏": (2616000, 3002166), "陕西": (36050000, 37327378),
    "甘肃": (25620000, 25575254), "青海": (5182000, 5626722), "宁夏": (5616000, 6301350),
    "新疆": (19250000, 21813334),
}
REGIONS_SORTED = sorted(CENSUS)
TOTAL_31_2000 = sum(p5 for p5, _ in CENSUS.values())          # 1262271800
TOTAL_31_2010 = sum(p6 for _, p6 in CENSUS.values())          # 1332774867
TOTAL_OFFICIAL_2000 = 1295330000
TOTAL_OFFICIAL_2010 = 1339724852
BIG_REGIONS = [name for name, (_, p6) in CENSUS.items() if p6 > 50000000]
GROWTH_DESC = sorted(
    [[name, CENSUS[name][1], CENSUS[name][1] - CENSUS[name][0]] for name in BIG_REGIONS],
    key=lambda item: -item[2])
EXPECTED_AVG_BIG = round(sum(item[1] for item in GROWTH_DESC) / len(GROWTH_DESC), 1)


def emit():
    print("AGENTBENCH_METRICS=" + json.dumps({"metrics": metrics, "evidence": evidence}))


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def close(a, b, tol=0.011):
    x, y = num(a), num(b)
    return x is not None and y is not None and abs(x - y) <= tol


def change_31(name):
    p5, p6 = CENSUS[name]
    return p6 / TOTAL_31_2010 * 100 - p5 / TOTAL_31_2000 * 100


def change_official(name):
    p5, p6 = CENSUS[name]
    return p6 / TOTAL_OFFICIAL_2010 * 100 - p5 / TOTAL_OFFICIAL_2000 * 100


try:
    import openpyxl
except Exception as exc:  # noqa: BLE001
    evidence["error"] = f"openpyxl 不可用: {str(exc)[:200]}"
    emit()
    raise SystemExit(0) from None

try:
    wbf = openpyxl.load_workbook("Excel.xlsx", data_only=False)
    wbv = openpyxl.load_workbook("Excel.xlsx", data_only=True)
except Exception as exc:  # noqa: BLE001
    evidence["error"] = f"Excel.xlsx 缺失或无法打开: {str(exc)[:200]}"
    emit()
    raise SystemExit(0) from None


def pick(workbook, keyword):
    for sheet in workbook.worksheets:
        if keyword in sheet.title:
            return sheet
    return None


def sheet_numbers(sheet):
    if sheet is None:
        return []
    return [num(cell.value) for row in sheet.iter_rows() for cell in row]


s5_f = pick(wbf, "第五次")
s5_v = pick(wbv, "第五次")
s6_f = pick(wbf, "第六次")
s6_v = pick(wbv, "第六次")
compare = pick(wbv, "比较")
stats_sheet = pick(wbv, "统计指标")
pivot = pick(wbv, "透视")

# ------------------------------------------------ e1 两表导入与表格样式
try:
    score = 0.0
    detail = []
    for sheet_f, _sheet_v in ((s5_f, s5_v), (s6_f, s6_v)):
        if sheet_f is None:
            detail.append("缺表")
            continue
        score += 10.0
        if not sheet_f.tables:
            detail.append(f"{sheet_f.title}:无表格对象")
            continue
        score += 20.0
        table = next(iter(sheet_f.tables.values()))
        style_info = table.tableStyleInfo
        striped = bool(style_info is not None and style_info.showRowStripes)
        if not striped:
            fills = [cell.fill for row_idx in (3, 5) for cell in sheet_f[row_idx]]
            striped = any(
                fill.fill_type == "solid"
                and getattr(fill.start_color, "rgb", None) not in (None, "00000000")
                for fill in fills)
        if striped:
            score += 20.0
        detail.append(f"{sheet_f.title}:条纹={striped}")
    metrics["e1"] = round(score, 2)
    evidence["e1"] = "; ".join(detail)
except Exception as exc:  # noqa: BLE001
    evidence["e1"] = str(exc)[:200]

# ------------------------------------------------ e2 千分位格式
try:
    ratios = []
    for sheet_f in (s5_f, s6_f):
        if sheet_f is None:
            ratios.append(0.0)
            continue
        hits = sum(
            1 for row in range(2, 33)
            if "#,##0" in str(sheet_f.cell(row=row, column=2).number_format))
        ratios.append(hits / 31)
    metrics["e2"] = round(100.0 * sum(ratios) / len(ratios), 2) if ratios else 0.0
    evidence["e2"] = f"千分位比例 {[round(r, 2) for r in ratios]}"
except Exception as exc:  # noqa: BLE001
    evidence["e2"] = str(exc)[:200]


def compare_header_columns():
    """返回 比较数据 表 {表头关键字: 列号} 与数据起始行。"""
    if compare is None:
        return None, None
    for row in range(1, min(compare.max_row, 6) + 1):
        headers = {}
        for cell in compare[row]:
            text = str(cell.value or "").strip()
            if text:
                headers[text] = cell.column
        if any("地区" in text for text in headers):
            return headers, row + 1
    return None, None


headers, first_row = compare_header_columns()


def find_column(*keywords):
    if not headers:
        return None
    for keyword in keywords:
        for text, column in headers.items():
            if keyword in text:
                return column
    return None


# ------------------------------------------------ e3 比较数据合并排序
try:
    if compare is None or first_row is None:
        evidence["e3"] = "缺少 比较数据 工作表或表头"
    else:
        column_a = find_column("地区")
        regions = [str(compare.cell(row=first_row + i, column=column_a).value or "").strip()
                   for i in range(31)]
        matched = sum(1 for actual, expected in zip(regions, REGIONS_SORTED, strict=False)
                      if actual == expected)
        metrics["e3"] = round(100.0 * matched / len(REGIONS_SORTED), 2)
        evidence["e3"] = f"{matched}/31 地区顺序正确"
except Exception as exc:  # noqa: BLE001
    evidence["e3"] = str(exc)[:200]


def region_value_map(column):
    """比较数据表中 {地区: 指定列数值}。"""
    result = {}
    if compare is None or first_row is None or column is None:
        return result
    region_col = find_column("地区")
    for row in range(first_row, compare.max_row + 1):
        region = str(compare.cell(row=row, column=region_col).value or "").strip()
        if region in CENSUS:
            result[region] = compare.cell(row=row, column=column).value
    return result


# ------------------------------------------------ e4 人口增长数列
try:
    growth_col = find_column("增长")
    values = region_value_map(growth_col)
    matched = sum(
        1 for name in CENSUS
        if close(values.get(name), CENSUS[name][1] - CENSUS[name][0], tol=0.5))
    metrics["e4"] = round(100.0 * matched / len(CENSUS), 2)
    evidence["e4"] = f"{matched}/31 增长数正确"
except Exception as exc:  # noqa: BLE001
    evidence["e4"] = str(exc)[:200]

# ------------------------------------------------ e5 比重变化列（两种合计基数口径均接受）
try:
    ratio_col = find_column("比重变化")
    values = region_value_map(ratio_col)
    matched = 0
    for name in CENSUS:
        value = num(values.get(name))
        if value is None:
            continue
        if abs(value - change_31(name)) <= 0.011 or abs(value - change_official(name)) <= 0.011:
            matched += 1
    metrics["e5"] = round(100.0 * matched / len(CENSUS), 2)
    evidence["e5"] = f"{matched}/31 比重变化正确"
except Exception as exc:  # noqa: BLE001
    evidence["e5"] = str(exc)[:200]

# ------------------------------------------------ e6 统计指标表
try:
    values = sheet_numbers(stats_sheet) if stats_sheet is not None else []
    if not values:
        values = [value for sheet in wbv.worksheets for value in sheet_numbers(sheet)]
    checks = [
        any(value is not None and abs(value - 31) < 0.5 for value in values),
        any(value is not None and abs(value - 10) < 0.5 for value in values),
        any(close(value, EXPECTED_AVG_BIG, tol=0.5) for value in values),
        any(close(value, TOTAL_OFFICIAL_2000, tol=0.5) for value in values)
        or any(close(value, TOTAL_31_2000, tol=0.5) for value in values),
        any(close(value, TOTAL_OFFICIAL_2010, tol=0.5) for value in values)
        or any(close(value, TOTAL_31_2010, tol=0.5) for value in values),
    ]
    metrics["e6"] = 100.0 * sum(checks) / len(checks)
    evidence["e6"] = f"五项指标 {sum(checks)}/5"
except Exception as exc:  # noqa: BLE001
    evidence["e6"] = str(exc)[:200]

# ------------------------------------------------ e7 超5000万地区降序榜
try:
    if pivot is None:
        evidence["e7"] = "缺少含 透视 的工作表"
    else:
        expected_order = [item[0] for item in GROWTH_DESC]
        found = []
        row_values = {}
        for row in pivot.iter_rows():
            region = next(
                (str(cell.value).strip() for cell in row
                 if str(cell.value or "").strip() in CENSUS), None)
            if region is None:
                continue
            if region in BIG_REGIONS and region not in found:
                found.append(region)
                row_values[region] = [num(cell.value) for cell in row]
        order_ok = found == expected_order
        value_hits = 0
        for name, pop, growth in GROWTH_DESC:
            values = row_values.get(name, [])
            if any(value is not None and abs(value - pop) < 0.5 for value in values) \
                    and any(value is not None and abs(value - growth) < 0.5 for value in values):
                value_hits += 1
        metrics["e7"] = round(50.0 * order_ok + 50.0 * value_hits / len(GROWTH_DESC), 2)
        evidence["e7"] = f"顺序ok={order_ok}, 数值{value_hits}/10"
except Exception as exc:  # noqa: BLE001
    evidence["e7"] = str(exc)[:200]

# ------------------------------------------------ e8 扁平汇总（透视降级）
try:
    values = sheet_numbers(pivot) if pivot is not None else []
    if not values:
        values = [value for sheet in wbv.worksheets for value in sheet_numbers(sheet)]
    year_hits = 0
    if any(close(value, TOTAL_31_2000, tol=0.5) for value in values) \
            or any(close(value, TOTAL_OFFICIAL_2000, tol=0.5) for value in values):
        year_hits += 1
    if any(close(value, TOTAL_31_2010, tol=0.5) for value in values) \
            or any(close(value, TOTAL_OFFICIAL_2010, tol=0.5) for value in values):
        year_hits += 1
    metrics["e8"] = 100.0 if year_hits == 2 else 50.0 * year_hits
    evidence["e8"] = f"全国合计命中 {year_hits}/2"
except Exception as exc:  # noqa: BLE001
    evidence["e8"] = str(exc)[:200]

emit()
