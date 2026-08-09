# NCRE Excel 操作题私有判分脚本（经典题库第2套：学生成绩单统计，command_metrics 协议）
# 用 openpyxl 检查 Excel.xlsx：表结构、数字格式、条件格式、SUM/AVERAGE、班级提取、
# 分类汇总与柱状图。公式写法不限，以单元格值为准；产物缺失时全 0。
import contextlib
import json
import re
import sys
import zipfile

with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

metrics = {f"e{i}": 0.0 for i in range(1, 9)}
evidence = {}

SUBJECTS = ["语文", "数学", "英语", "物理", "化学", "生物", "政治"]
# (学号, 姓名, 语文, 数学, 英语, 物理, 化学, 生物, 政治)
GRADES = {
    "12010101": ("陈晓", [118, 125, 132, 88, 92, 78, 85]),
    "12010102": ("李倩", [105, 98, 110, 76, 84, 70, 80]),
    "12010201": ("王涛", [96, 112, 108, 82, 78, 74, 77]),
    "12010202": ("刘倩", [110, 135, 121, 91, 88, 82, 84]),
    "12010301": ("张悦", [88, 95, 102, 72, 80, 68, 75]),
    "12010302": ("赵磊", [101, 120, 115, 85, 90, 79, 82]),
    "12020101": ("孙悦", [112, 108, 99, 80, 86, 75, 79]),
    "12020102": ("周航", [95, 130, 125, 89, 93, 81, 86]),
    "12020201": ("吴迪", [108, 102, 95, 77, 82, 71, 78]),
    "12020301": ("郑爽", [92, 116, 106, 84, 87, 76, 81]),
    "12020302": ("冯军", [115, 128, 118, 90, 95, 83, 88]),
    "12030101": ("蒋丽", [99, 90, 112, 74, 79, 66, 73]),
}
EXPECTED_AVG = {
    "12010101": 102.57, "12010102": 89.0, "12010201": 89.57, "12010202": 101.57,
    "12010301": 82.86, "12010302": 96.0, "12020101": 91.29, "12020102": 99.86,
    "12020201": 87.57, "12020301": 91.71, "12020302": 102.43, "12030101": 84.71,
}
EXPECTED_CLASS_AVG = {
    "1班": [103.0, 114.17, 114.67, 82.33, 85.33, 75.17, 80.5],
    "2班": [104.4, 116.8, 108.6, 84.0, 88.6, 77.2, 82.4],
    "3班": [99.0, 90.0, 112.0, 74.0, 79.0, 66.0, 73.0],
}
DATA_ROWS = list(range(2, 14))


def emit():
    print("AGENTBENCH_METRICS=" + json.dumps({"metrics": metrics, "evidence": evidence}))


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def close(a, b, tol=0.011):
    x, y = num(a), num(b)
    return x is not None and y is not None and abs(x - y) <= tol


try:
    import openpyxl
except Exception as exc:  # noqa: BLE001
    evidence["error"] = f"openpyxl 不可用: {str(exc)[:200]}"
    emit()
    raise SystemExit(0) from None

try:
    wbf = openpyxl.load_workbook("Excel.xlsx", data_only=False)
    wbv = openpyxl.load_workbook("Excel.xlsx", data_only=True)
except Exception as exc:  # noqa: BLE001
    evidence["error"] = f"Excel.xlsx 缺失或无法打开: {str(exc)[:200]}"
    emit()
    raise SystemExit(0) from None


def pick(workbook, keyword, fallback_index):
    for sheet in workbook.worksheets:
        if keyword in sheet.title:
            return sheet
    return workbook.worksheets[fallback_index]


ws_f = pick(wbf, "成绩", 0)
ws_v = pick(wbv, "成绩", 0)

# ------------------------------------------------ e1 表结构与原始数据
try:
    header_ok = (str(ws_f["A1"].value).strip() == "学号"
                 and str(ws_f["B1"].value).strip() == "姓名"
                 and [str(ws_f.cell(row=1, column=c).value).strip()
                      for c in range(4, 11)] == SUBJECTS)
    row_hits = 0
    for row in DATA_ROWS:
        student_id = str(ws_v.cell(row=row, column=1).value or "").strip()
        info = GRADES.get(student_id)
        if info is None:
            continue
        name_ok = str(ws_v.cell(row=row, column=2).value or "").strip() == info[0]
        scores_ok = all(
            close(ws_v.cell(row=row, column=4 + i).value, score)
            for i, score in enumerate(info[1]))
        if name_ok and scores_ok:
            row_hits += 1
    metrics["e1"] = round(40.0 * header_ok + 60.0 * row_hits / len(DATA_ROWS), 2)
    evidence["e1"] = f"表头ok={header_ok}, 数据行{row_hits}/12"
except Exception as exc:  # noqa: BLE001
    evidence["e1"] = str(exc)[:200]

# ------------------------------------------------ e2 数字格式：学号文本、成绩两位小数
try:
    id_hits = sum(
        1 for row in DATA_ROWS
        if "@" in str(ws_f.cell(row=row, column=1).number_format)
        or ws_f.cell(row=row, column=1).data_type == "s"
    )
    score_cells = 0
    score_hits = 0
    for row in DATA_ROWS:
        for column in list(range(4, 11)) + [12]:
            score_cells += 1
            if "0.00" in str(ws_f.cell(row=row, column=column).number_format):
                score_hits += 1
    metrics["e2"] = round(50.0 * id_hits / len(DATA_ROWS)
                          + 50.0 * score_hits / score_cells, 2)
    evidence["e2"] = f"学号文本{id_hits}/12, 0.00格式{score_hits}/{score_cells}"
except Exception as exc:  # noqa: BLE001
    evidence["e2"] = str(exc)[:200]

# ------------------------------------------------ e3 条件格式
try:
    column_rules = {}
    for cf in ws_f.conditional_formatting:
        for rule in cf.rules:
            for rng in cf.sqref.ranges:
                for column in range(rng.min_col, rng.max_col + 1):
                    column_rules.setdefault(column, []).append(rule)

    def rule_matches(column, operator, threshold, want_fill):
        for rule in column_rules.get(column, []):
            formulas = [num(item) for item in getattr(rule, "formula", []) or []]
            if getattr(rule, "operator", None) != operator:
                continue
            if not any(value is not None and abs(value - threshold) < 0.001 for value in formulas):
                continue
            dxf = getattr(rule, "dxf", None)
            if dxf is None:
                continue
            if want_fill and getattr(dxf, "fill", None) is not None:
                return True
            if not want_fill:
                font = getattr(dxf, "font", None)
                if font is not None and getattr(font, "color", None) is not None:
                    return True
        return False

    checks = [rule_matches(column, "greaterThanOrEqual", 110, True) for column in (4, 5, 6)]
    checks += [rule_matches(column, "greaterThan", 95, False) for column in (7, 8, 9, 10)]
    metrics["e3"] = round(100.0 * sum(checks) / len(checks), 2)
    evidence["e3"] = f"{sum(checks)}/7 列条件格式正确"
except Exception as exc:  # noqa: BLE001
    evidence["e3"] = str(exc)[:200]

# ------------------------------------------------ e4 总分(SUM)与平均分(AVERAGE)
try:
    total_hits = avg_hits = 0
    for row in DATA_ROWS:
        student_id = str(ws_v.cell(row=row, column=1).value or "").strip()
        expected_scores = GRADES.get(student_id, (None, [None] * 7))[1]
        cell_scores = [num(ws_v.cell(row=row, column=c).value) for c in range(4, 11)]
        scores = [a if a is not None else b for a, b in zip(cell_scores, expected_scores, strict=False)]
        # K列 总分
        formula = ws_f.cell(row=row, column=11).value
        formula_ok = isinstance(formula, str) and "SUM" in formula.upper()
        value = num(ws_v.cell(row=row, column=11).value)
        if value is None and formula_ok and all(s is not None for s in scores):
            value = sum(scores)
        if formula_ok and value is not None and all(s is not None for s in scores) \
                and close(value, sum(scores)):
            total_hits += 1
        # L列 平均分
        formula = ws_f.cell(row=row, column=12).value
        formula_ok = isinstance(formula, str) and "AVERAGE" in formula.upper()
        value = num(ws_v.cell(row=row, column=12).value)
        expected_avg = EXPECTED_AVG.get(student_id)
        if value is None and formula_ok and expected_avg is not None:
            value = expected_avg
        if formula_ok and expected_avg is not None and close(value, expected_avg):
            avg_hits += 1
    metrics["e4"] = round(50.0 * total_hits / len(DATA_ROWS)
                          + 50.0 * avg_hits / len(DATA_ROWS), 2)
    evidence["e4"] = f"总分{total_hits}/12, 平均{avg_hits}/12"
except Exception as exc:  # noqa: BLE001
    evidence["e4"] = str(exc)[:200]

# ------------------------------------------------ e5 班级提取(MID)
try:
    hits = 0
    for row in DATA_ROWS:
        student_id = str(ws_v.cell(row=row, column=1).value or "").strip()
        expected_class = f"{int(student_id[2:4])}班" if len(student_id) >= 4 else None
        formula = str(ws_f.cell(row=row, column=3).value or "")
        formula_ok = "MID" in formula.upper()
        cached = ws_v.cell(row=row, column=3).value
        value_ok = cached is not None and str(cached).strip() == expected_class
        if not value_ok and formula_ok and expected_class:
            match = re.search(
                rf"MID\(\s*\$?A{row}\s*,\s*(\d+)\s*,\s*(\d+)\s*\)", formula, re.IGNORECASE)
            if match:
                start, length = int(match.group(1)), int(match.group(2))
                code = student_id[start - 1:start - 1 + length]
                value_ok = code.isdigit() and f"{int(code)}班" == expected_class
        if formula_ok and value_ok:
            hits += 1
    metrics["e5"] = round(100.0 * hits / len(DATA_ROWS), 2)
    evidence["e5"] = f"{hits}/12 班级正确"
except Exception as exc:  # noqa: BLE001
    evidence["e5"] = str(exc)[:200]

# ------------------------------------------------ e6 分类汇总工作表
try:
    summary_f = next((ws for ws in wbf.worksheets if "分类汇总" in ws.title), None)
    summary_v = next((ws for ws in wbv.worksheets if "分类汇总" in ws.title), None)
    if summary_f is None:
        evidence["e6"] = "缺少含 分类汇总 的工作表"
    else:
        score = 20.0
        tab_color = summary_f.sheet_properties.tabColor
        if tab_color is not None and (
                getattr(tab_color, "rgb", None) not in (None, "00000000")
                or getattr(tab_color, "theme", None) is not None):
            score += 20.0
        values = [num(cell.value) for row in summary_v.iter_rows() for cell in row]
        for _class_name, expected in EXPECTED_CLASS_AVG.items():
            matched = sum(
                1 for want in expected
                if any(value is not None and abs(value - want) <= 0.011 for value in values))
            score += 20.0 * matched / len(expected)
        metrics["e6"] = round(score, 2)
        evidence["e6"] = f"tabColor={tab_color is not None}, score={score:.0f}"
except Exception as exc:  # noqa: BLE001
    evidence["e6"] = str(exc)[:200]

# ------------------------------------------------ e7 柱状分析图工作表
try:
    metrics["e7"] = 100.0 if any("柱状分析图" in ws.title for ws in wbf.worksheets) else 0.0
    evidence["e7"] = ",".join(ws.title for ws in wbf.worksheets)
except Exception as exc:  # noqa: BLE001
    evidence["e7"] = str(exc)[:200]

# ------------------------------------------------ e8 簇状柱形图
try:
    with zipfile.ZipFile("Excel.xlsx") as archive:
        chart_xml = "".join(
            archive.read(name).decode("utf-8", errors="replace")
            for name in archive.namelist()
            if re.search(r"xl/charts/chart\d+\.xml$", name)
        )
    if "<c:barChart" not in chart_xml and "<barChart" not in chart_xml:
        evidence["e8"] = "无 barChart"
    else:
        score = 40.0 if re.search(r'<(?:c:)?barDir val="col"', chart_xml) else 0.0
        series_count = chart_xml.count("<c:ser>") + chart_xml.count("<ser>")
        score += 30.0 if series_count == 3 else 0.0
        # 系列标题/分类可能是单元格引用（strRef），回查工作簿缓存值后再匹配科目名
        haystack = chart_xml
        for match in re.finditer(
                r"'([^']+)'!(\$?[A-Z]{1,3}\$?\d+)(:\$?[A-Z]{1,3}\$?\d+)?", chart_xml):
            sheet_name = match.group(1)
            if sheet_name not in wbf.sheetnames:
                continue
            region = wbf[sheet_name][match.group(2) + (match.group(3) or "")]
            rows = region if isinstance(region, tuple) else ((region,),)
            for row_cells in rows:
                for cell in row_cells:
                    haystack += " " + str(cell.value or "")
        subject_hits = sum(1 for subject in SUBJECTS if subject in haystack)
        score += 30.0 * subject_hits / len(SUBJECTS)
        metrics["e8"] = round(score, 2)
        evidence["e8"] = f"系列数{series_count}, 科目{subject_hits}/7"
except Exception as exc:  # noqa: BLE001
    evidence["e8"] = str(exc)[:200]

emit()
