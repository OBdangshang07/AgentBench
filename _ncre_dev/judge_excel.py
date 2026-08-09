# NCRE Excel 操作题私有判分脚本（AgentBench command_metrics 协议）
# 用 openpyxl 检查 Excel.xlsx：表对象、会计格式、VLOOKUP、小计与统计单元格。
# 公式写法不限，以单元格计算值为准；无缓存值时尝试轻量公式求值。产物缺失时全 0。
import contextlib
import datetime as dt
import json
import re
import sys

with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

metrics = {f"e{i}": 0.0 for i in range(1, 10)}
evidence = {}

EXPECTED = {"B3": 11925.0, "B4": 2385.0, "B5": 1140.0, "B6": 252.25}
BOOK = {
    "BK-83021": ("《MS Office高级应用》", 45.00),
    "BK-10001": ("《计算机基础教程》", 30.00),
    "BK-10002": ("《Word实用技巧》", 32.00),
    "BK-10003": ("《Excel数据分析》", 39.00),
    "BK-10004": ("《PowerPoint演示设计》", 35.00),
    "BK-10005": ("《数据库技术基础》", 42.00),
    "BK-10006": ("《计算机网络技术基础》", 38.00),
    "BK-10007": ("《Photoshop图像处理》", 49.00),
}


def emit():
    print("AGENTBENCH_METRICS=" + json.dumps({"metrics": metrics, "evidence": evidence}))


def num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def close(a, b, tol=0.01):
    x, y = num(a), num(b)
    return x is not None and y is not None and abs(x - y) <= tol


def parse_date(text):
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return dt.datetime.strptime(str(text).strip(), fmt)
        except ValueError:
            continue
    return None


try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except Exception as exc:  # noqa: BLE001
    evidence["error"] = f"openpyxl 不可用: {str(exc)[:200]}"
    emit()
    raise SystemExit(0) from None

wbf = wbv = orders_f = orders_v = ref_sheet = report = None
header_row = 2
columns = {}
data_rows = []
table_info = None

try:
    wbf = openpyxl.load_workbook("Excel.xlsx", data_only=False)
    wbv = openpyxl.load_workbook("Excel.xlsx", data_only=True)

    def pick(workbook, keyword, fallback_index):
        for sheet in workbook.worksheets:
            if keyword in sheet.title:
                return sheet
        return workbook.worksheets[fallback_index]

    orders_f = pick(wbf, "订单明细", 0)
    orders_v = pick(wbv, "订单明细", 0)
    ref_sheet = pick(wbf, "编号对照", min(1, len(wbf.worksheets) - 1))
    report = pick(wbv, "统计报告", len(wbv.worksheets) - 1)
    report_f = pick(wbf, "统计报告", len(wbf.worksheets) - 1)

    for row in range(1, 4):
        values = [(cell.column, str(cell.value).strip())
                  for cell in orders_f[row] if cell.value is not None]
        if any(name == "订单编号" for _, name in values):
            header_row = row
            columns = {name: col for col, name in values}
            break
    if columns:
        id_col = columns.get("订单编号")
        for row in range(header_row + 1, orders_f.max_row + 1):
            if orders_f.cell(row=row, column=id_col).value not in (None, ""):
                data_rows.append(row)
    for table in orders_f.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        table_info = {
            "ref": table.ref,
            "min_row": min_row, "max_row": max_row,
            "headers": {
                str(orders_f.cell(row=min_row, column=c).value).strip(): c
                for c in range(min_col, max_col + 1)
            },
        }
        break
except Exception as exc:  # noqa: BLE001
    evidence["error"] = f"Excel.xlsx 读取失败: {str(exc)[:200]}"
    emit()
    raise SystemExit(0) from None


def resolve_structured(token, row):
    match = re.match(r"\[@(.+?)\]$", token.strip())
    if not match or table_info is None:
        return None
    col = table_info["headers"].get(match.group(1).strip())
    if col is None:
        return None
    return orders_v.cell(row=row, column=col).value


def resolve_vlookup(formula, row):
    match = re.search(
        r"VLOOKUP\(\s*([^,()]+?)\s*,\s*[^,]+?\s*,\s*(\d+)\s*,\s*(0|FALSE|FALSE\(\))\s*\)",
        formula, re.IGNORECASE)
    if not match:
        return None
    lookup_token, column_index = match.group(1).strip(), int(match.group(2))
    if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", lookup_token):
        lookup = orders_v[lookup_token.replace("$", "")].value
    else:
        lookup = resolve_structured(lookup_token, row)
    if lookup is None:
        return None
    info = BOOK.get(str(lookup).strip())
    if info is None:
        return None
    return info[0] if column_index == 2 else info[1]


def compare_criterion(value, criterion):
    op_match = re.match(r"(>=|<=|<>|>|<|=)(.*)$", str(criterion).strip())
    op, rhs = (op_match.group(1), op_match.group(2)) if op_match else ("=", str(criterion))
    rhs = rhs.strip().strip('"')
    date_rhs = parse_date(rhs)
    if isinstance(value, dt.datetime) or date_rhs is not None:
        left = value if isinstance(value, dt.datetime) else parse_date(value)
        if left is None or date_rhs is None:
            return False
        return {"=": left == date_rhs, ">": left > date_rhs, "<": left < date_rhs,
                ">=": left >= date_rhs, "<=": left <= date_rhs, "<>": left != date_rhs}[op]
    rhs_num = num(rhs)
    if rhs_num is not None:
        left_num = num(value)
        if left_num is None:
            return False
        return {"=": close(left_num, rhs_num), ">": left_num > rhs_num,
                "<": left_num < rhs_num, ">=": left_num >= rhs_num,
                "<=": left_num <= rhs_num, "<>": not close(left_num, rhs_num)}[op]
    text = str(value)
    if "*" in rhs or "?" in rhs:
        pattern = "^" + re.escape(rhs).replace(r"\*", ".*").replace(r"\?", ".") + "$"
        matched = re.match(pattern, text) is not None
    else:
        matched = text == rhs
    return matched if op == "=" else not matched


def resolve_sheet_cell(row, col, depth=0):
    """先取缓存值，缺失时递归求值公式（支持 VLOOKUP/乘法/单元格引用链）。"""
    cached = orders_v.cell(row=row, column=col).value
    if cached is not None:
        return cached
    if depth > 8:
        return None
    formula = orders_f.cell(row=row, column=col).value
    if not (isinstance(formula, str) and formula.startswith("=")):
        return None
    return eval_cell_formula(formula[1:], row, depth)


def resolve_token(token, row, depth=0):
    token = token.strip()
    resolved = resolve_structured(token, row)
    if resolved is not None:
        return resolved
    if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", token):
        cell_ref = token.replace("$", "")
        match = re.match(r"([A-Z]+)(\d+)", cell_ref)
        from openpyxl.utils import column_index_from_string
        return resolve_sheet_cell(
            int(match.group(2)), column_index_from_string(match.group(1)), depth + 1)
    return num(token)


def eval_cell_formula(body, row, depth=0):
    body = body.strip()
    if depth > 8:
        return None
    vlookup = resolve_vlookup(body, row)
    if vlookup is not None:
        return vlookup
    product = re.fullmatch(r"([^*]+?)\*([^*]+?)", body)
    if product:
        factors = [resolve_token(token, row, depth + 1) for token in product.groups()]
        if all(num(f) is not None for f in factors):
            return num(factors[0]) * num(factors[1])
    return resolve_token(body, row, depth + 1)


def range_cells(spec):
    spec = spec.split("!")[-1].replace("$", "")
    if ":" not in spec:
        return [resolve_token(spec, header_row + 1)]
    start, end = spec.split(":")
    min_col, min_row, max_col, max_row = range_boundaries(f"{start}:{end}")
    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cells.append(resolve_sheet_cell(r, c))
    return cells


def sumifs(args):
    if len(args) < 3 or len(args) % 2 == 0:
        return None
    total, pairs = 0.0, args[1:]
    values = range_cells(args[0])
    for index in range(0, len(pairs), 2):
        crit_values = range_cells(pairs[index])
        if len(crit_values) != len(values):
            return None
        for i, value in enumerate(values):
            if value is None or not compare_criterion(value, pairs[index + 1]):
                values[i] = None
    for value in values:
        if value is not None and num(value) is not None:
            total += num(value)
    return total


def eval_stats(formula):
    body = formula.strip().lstrip("=").strip()
    match = re.fullmatch(r"SUM\(([^()]*)\)(\s*[-+*/]\s*[\d.]+)?", body, re.IGNORECASE)
    if match:
        total = 0.0
        for part in match.group(1).split(","):
            for value in range_cells(part.strip()):
                total += num(value) or 0.0
        return apply_trailing(total, match.group(2))
    match = re.fullmatch(r"SUMIFS\((.+)\)(\s*[-+*/]\s*[\d.]+)?", body, re.IGNORECASE)
    if match:
        args = [item.strip() for item in match.group(1).split(",")]
        result = sumifs(args)
        return None if result is None else apply_trailing(result, match.group(2))
    match = re.fullmatch(r"SUMPRODUCT\((.+)\)(\s*[-+*/]\s*[\d.]+)?", body, re.IGNORECASE)
    if match:
        inner = match.group(1)
        arrays = []
        for factor in re.split(r"\)\s*\*\s*\(", inner.strip("()")):
            factor = factor.strip("() ")
            compare = re.match(r"(.+?)(>=|<=|<>|>|<|=)(.+)$", factor)
            if compare:
                arrays.append([compare_criterion(v, compare.group(2) + compare.group(3))
                               for v in range_cells(compare.group(1))])
            else:
                arrays.append([num(v) or 0.0 for v in range_cells(factor)])
        if not arrays:
            return None
        length = len(arrays[0])
        if any(len(a) != length for a in arrays):
            return None
        total = 0.0
        for i in range(length):
            product = 1.0
            for arr in arrays:
                item = arr[i]
                if isinstance(item, bool):
                    product *= 1.0 if item else 0.0
                else:
                    product *= num(item) or 0.0
            total += product
        return apply_trailing(total, match.group(2))
    value = num(body)
    return value


def apply_trailing(total, suffix):
    if not suffix or total is None:
        return total
    operator = suffix.strip()[0]
    operand = num(suffix.strip()[1:])
    if operand is None:
        return None
    if operator == "/":
        return total / operand if operand else None
    if operator == "*":
        return total * operand
    if operator == "+":
        return total + operand
    return total - operand


# ------------------------------------------------ e1 套用表格格式
try:
    ok = False
    for table in orders_f.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row <= header_row and data_rows and max_row >= max(data_rows) \
                and (max_col - min_col + 1) >= 6:
            ok = True
    metrics["e1"] = 100.0 if ok else 0.0
    evidence["e1"] = ",".join(t.ref for t in orders_f.tables.values()) or "无表对象"
except Exception as exc:  # noqa: BLE001
    evidence["e1"] = str(exc)[:200]

# ------------------------------------------------ e2 会计专用格式
try:
    ratios = []
    for column_name in ("单价", "小计"):
        col = columns.get(column_name)
        if col is None:
            ratios.append(0.0)
            continue
        hits = sum(
            1 for row in data_rows
            if "￥" in str(orders_f.cell(row=row, column=col).number_format)
        )
        ratios.append(hits / len(data_rows) if data_rows else 0.0)
    metrics["e2"] = round(100.0 * sum(ratios) / len(ratios), 2)
except Exception as exc:  # noqa: BLE001
    evidence["e2"] = str(exc)[:200]

# ------------------------------------------------ e3/e4 VLOOKUP
def judge_vlookup(metric_key, column_name, expect_index):
    col = columns.get(column_name)
    if col is None or not data_rows:
        evidence[metric_key] = f"缺少列 {column_name}"
        return
    ok_rows = 0
    for row in data_rows:
        formula = orders_f.cell(row=row, column=col).value
        formula_ok = isinstance(formula, str) and "VLOOKUP" in formula.upper()
        expected = BOOK.get(
            str(orders_v.cell(row=row, column=columns["图书编号"]).value).strip()
        )
        if expected is None:
            continue
        expected_value = expected[expect_index]
        cached = orders_v.cell(row=row, column=col).value
        if cached is not None:
            value_ok = (str(cached).strip() == str(expected_value)) \
                if expect_index == 0 else close(cached, expected_value)
        else:
            resolved = resolve_vlookup(str(formula or ""), row)
            value_ok = (resolved is not None and (
                str(resolved).strip() == str(expected_value)
                if expect_index == 0 else close(resolved, expected_value)))
        if formula_ok and value_ok:
            ok_rows += 1
    metrics[metric_key] = round(100.0 * ok_rows / len(data_rows), 2)


try:
    judge_vlookup("e3", "图书名称", 0)
    judge_vlookup("e4", "单价", 1)
except Exception as exc:  # noqa: BLE001
    evidence["e3/e4"] = str(exc)[:200]

# ------------------------------------------------ e5 小计计算
try:
    price_col = columns.get("单价")
    qty_col = columns.get("销量(本)") or columns.get("销量")
    sub_col = columns.get("小计")
    ok_rows = 0
    if price_col and qty_col and sub_col and data_rows:
        for row in data_rows:
            cell_value = orders_f.cell(row=row, column=sub_col).value
            cached = orders_v.cell(row=row, column=sub_col).value
            price = orders_v.cell(row=row, column=price_col).value
            if price is None:
                price = resolve_vlookup(
                    str(orders_f.cell(row=row, column=price_col).value or ""), row)
            qty = orders_v.cell(row=row, column=qty_col).value
            expected = num(price) * num(qty) if num(price) is not None and num(qty) is not None else None
            if expected is None:
                continue
            value_ok = close(cached, expected)
            if not value_ok and isinstance(cell_value, str) and cell_value.startswith("="):
                product = re.fullmatch(
                    r"=\s*([^*]+?)\s*\*\s*([^*]+?)\s*$", cell_value.strip())
                if product:
                    factors = []
                    for token in product.groups():
                        token = token.strip()
                        resolved = resolve_structured(token, row)
                        if resolved is None and re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", token):
                            cell_ref = token.replace("$", "")
                            resolved = orders_v[cell_ref].value
                            if resolved is None:
                                resolved = resolve_vlookup(
                                    str(orders_f[cell_ref].value or ""),
                                    int(re.search(r"\d+", cell_ref).group()))
                        factors.append(num(resolved))
                    if all(f is not None for f in factors):
                        value_ok = close(factors[0] * factors[1], expected)
            if isinstance(cell_value, str) and cell_value.startswith("=") and value_ok:
                ok_rows += 1
        metrics["e5"] = round(100.0 * ok_rows / len(data_rows), 2)
except Exception as exc:  # noqa: BLE001
    evidence["e5"] = str(exc)[:200]

# ------------------------------------------------ e6-e9 统计报告
for metric_key, address in (("e6", "B3"), ("e7", "B4"), ("e8", "B5"), ("e9", "B6")):
    try:
        expected = EXPECTED[address]
        cached = report[address].value
        formula = report_f[address].value
        resolved = num(cached)
        if resolved is None and isinstance(formula, str) and formula.startswith("="):
            resolved = num(eval_stats(formula))
        value_ok = resolved is not None and close(resolved, expected)
        score = 100.0 if value_ok else 0.0
        if metric_key == "e9":
            fmt = str(report_f[address].number_format)
            format_ok = "0.00" in fmt
            score = (80.0 if value_ok else 0.0) + (20.0 if format_ok else 0.0)
            evidence["e9"] = f"value={resolved}, fmt={fmt[:40]}"
        else:
            evidence[metric_key] = f"value={resolved}, expected={expected}"
        metrics[metric_key] = round(score, 2)
    except Exception as exc:  # noqa: BLE001
        evidence[metric_key] = str(exc)[:200]

emit()
